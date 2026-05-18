import functools
import hmac
import os
from io import BytesIO

from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from db import (
    cancel_order as cancel_order_record,
    clear_table as clear_table_record,
    get_dashboard_data,
    get_sales_export_data,
    get_menu_items,
    get_table_status as get_table_status_record,
    init_db,
    place_order as place_order_record,
    reset_operation_data as reset_operation_data_record,
    toggle_item_pay as toggle_item_pay_record,
    toggle_serve as toggle_serve_record,
    update_menu_item,
)

app = Flask(__name__)
# Cache static assets (menu images, CSS, JS) for 2 hours during event operation.
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 7200
app.secret_key = os.environ.get('SECRET_KEY', 'sommelier-dev-fallback-key-change-in-prod')
if not os.environ.get('SECRET_KEY'):
    import logging
    logging.getLogger(__name__).warning(
        'SECRET_KEY env var not set — using development fallback. '
        'Set SECRET_KEY before deploying to a shared network.'
    )
init_db()

_ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'tedchang')
_ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'changsikhi')


def _check_credentials(username: str, password: str) -> bool:
    user_ok = hmac.compare_digest(username.encode(), _ADMIN_USERNAME.encode())
    pass_ok = hmac.compare_digest(password.encode(), _ADMIN_PASSWORD.encode())
    return user_ok and pass_ok


def _admin_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            if request.path.startswith('/api/'):
                return jsonify({"status": "error", "message": "인증이 필요합니다."}), 401
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated


# ── request validation helpers ──────────────────────────────────────────
def _bad(msg: str):
    return jsonify({"status": "error", "message": msg}), 400


def _valid_order_id(value) -> bool:
    """True if value is a positive integer (not bool) suitable as order_id."""
    return isinstance(value, int) and not isinstance(value, bool) and value > 0


@app.route('/')
def home():
    return redirect(url_for('guest_menu'))

@app.route('/menu')
def guest_menu():
    return render_template('customer_menu.html', menus=get_menu_items(guest_only=True))

@app.route('/table/<table_id>')
def server_page(table_id):
    return render_template('server_order.html', menus=get_menu_items(), table_id=table_id)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('admin_logged_in'):
        return redirect(url_for('admin_panel'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        if _check_credentials(username, password):
            session['admin_logged_in'] = True
            next_url = request.args.get('next', '')
            safe = next_url.startswith('/') and not next_url.startswith('//')
            return redirect(next_url if safe else url_for('admin_panel'))
        error = '아이디 또는 비밀번호가 올바르지 않습니다.'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('login'))

@app.route('/admin')
@_admin_required
def admin_panel():
    return render_template('admin_panel.html', data=get_dashboard_data())

@app.route('/api/order', methods=['POST'])
def place_order():
    req = request.get_json(silent=True)
    if req is None:
        return _bad("요청 본문이 올바른 JSON 형식이어야 합니다.")
    table_id = str(req.get('table_id', '')).strip()
    if not table_id:
        return _bad("테이블 번호가 없거나 비어 있습니다.")
    items = req.get('items', [])
    if not isinstance(items, list) or not items:
        return _bad("주문 항목이 없습니다.")
    if any(not isinstance(item, str) or not item.strip() for item in items):
        return _bad("주문 항목 형식이 올바르지 않습니다.")
    try:
        place_order_record(table_id, items)
        return jsonify({"status": "success"})
    except ValueError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400

@app.route('/api/orders', methods=['GET'])
@_admin_required
def get_admin_data():
    return jsonify(get_dashboard_data())

@app.route('/api/admin/export_sales', methods=['GET'])
@_admin_required
def export_sales():
    export_data = get_sales_export_data()
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["항목", "값"])
    summary_sheet.append(["누적 매출 (원장 합산)", export_data["ledger_revenue"]])
    summary_sheet.append(["유효 판매 건수", export_data["order_count"]])
    summary_sheet.append(["취소 건수", export_data["cancelled_count"]])
    summary_sheet.append(["미입금 건수", export_data["unpaid_count"]])
    summary_sheet.append(["미입금 금액", export_data["unpaid_amount"]])

    item_sheet = workbook.create_sheet("Menu Sales")
    item_sheet.append(["메뉴명", "판매 수량", "매출 합계", "평균 판매가"])
    for row in export_data["menu_sales"]:
        item_sheet.append([row["menu_name"], row["quantity"], row["revenue"], row["avg_price"]])

    order_sheet = workbook.create_sheet("Order Ledger")
    order_sheet.append([
        "주문 ID",
        "주문 시각",
        "테이블",
        "메뉴명",
        "판매가",
        "서빙 상태",
        "입금 여부",
        "최종 상태",
        "정리 시각",
        "취소 시각",
    ])
    for row in export_data["orders"]:
        order_sheet.append([
            row["ledger_id"],
            row["created_at"],
            row["table_id"],
            row["menu_name"],
            row["price"],
            row["status"],
            "입금" if row["is_paid"] else "미입금",
            row["final_state"],
            row["cleared_at"] or "",
            row["cancelled_at"] or "",
        ])

    for sheet in workbook.worksheets:
        header_fill = PatternFill("solid", fgColor="862633")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max_length + 3, 36)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="sommelier_sales_ledger.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route('/api/table_status/<table_id>', methods=['GET'])
def get_table_status(table_id):
    return jsonify(get_table_status_record(str(table_id)))

@app.route('/api/admin/reset_operation_data', methods=['POST'])
@_admin_required
def reset_operation_data():
    req = request.get_json(silent=True)
    if req is None:
        return _bad("요청 본문이 올바른 JSON 형식이어야 합니다.")
    if req.get('confirm') != 'RESET':
        return _bad("확인 코드가 올바르지 않습니다. 'RESET'을 정확히 입력하세요.")
    reset_operation_data_record()
    return jsonify({"status": "success"})


@app.route('/api/admin/update_menu', methods=['POST'])
@_admin_required
def admin_update_menu():
    req = request.get_json(silent=True)
    if req is None:
        return _bad("요청 본문이 올바른 JSON 형식이어야 합니다.")
    try:
        menu_id = int(req['id'])
        price = int(req['price'])
        stock = int(req['stock'])
    except (KeyError, TypeError, ValueError):
        return _bad("id, price, stock 값이 올바르지 않습니다.")
    if menu_id <= 0:
        return _bad("메뉴 ID가 올바르지 않습니다.")
    if price < 0:
        return _bad("가격은 0 이상이어야 합니다.")
    if stock < 0:
        return _bad("재고는 0 이상이어야 합니다.")
    update_menu_item(menu_id, price, stock)
    return jsonify({"status": "success"})

@app.route('/api/item_pay', methods=['POST'])
@_admin_required
def toggle_item_pay():
    req = request.get_json(silent=True)
    if req is None:
        return _bad("요청 본문이 올바른 JSON 형식이어야 합니다.")
    t_id = str(req.get('table_id', '')).strip()
    if not t_id:
        return _bad("테이블 번호가 없거나 비어 있습니다.")
    oid = req.get('order_id')
    if not _valid_order_id(oid):
        return _bad("order_id가 올바르지 않습니다.")
    if toggle_item_pay_record(t_id, oid):
        return jsonify({"status": "success"})
    return jsonify({"status": "fail"}), 400

@app.route('/api/serve', methods=['POST'])
def toggle_serve():
    req = request.get_json(silent=True)
    if req is None:
        return _bad("요청 본문이 올바른 JSON 형식이어야 합니다.")
    t_id = str(req.get('table_id', '')).strip()
    if not t_id:
        return _bad("테이블 번호가 없거나 비어 있습니다.")
    oid = req.get('order_id')
    if not _valid_order_id(oid):
        return _bad("order_id가 올바르지 않습니다.")
    if toggle_serve_record(t_id, oid):
        return jsonify({"status": "success"})
    return jsonify({"status": "fail"}), 400

@app.route('/api/cancel_order', methods=['POST'])
def cancel_order():
    req = request.get_json(silent=True)
    if req is None:
        return _bad("요청 본문이 올바른 JSON 형식이어야 합니다.")
    t_id = str(req.get('table_id', '')).strip()
    if not t_id:
        return _bad("테이블 번호가 없거나 비어 있습니다.")
    oid = req.get('order_id')
    if not _valid_order_id(oid):
        return _bad("order_id가 올바르지 않습니다.")
    if cancel_order_record(t_id, oid):
        return jsonify({"status": "success"})
    return jsonify({"status": "fail"}), 400

@app.route('/api/clear', methods=['POST'])
@_admin_required
def clear_table():
    req = request.get_json(silent=True)
    if req is None:
        return _bad("요청 본문이 올바른 JSON 형식이어야 합니다.")
    t_id = str(req.get('table_id', '')).strip()
    if not t_id:
        return _bad("테이블 번호가 없거나 비어 있습니다.")
    if clear_table_record(t_id):
        return jsonify({"status": "success"})
    return jsonify({"status": "fail"}), 400

@app.template_filter('format_price')
def format_price(value):
    return format(value, ',')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
